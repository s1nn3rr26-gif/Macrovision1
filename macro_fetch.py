# ─────────────────────────────────────────────────────────
#  MacroVision · macro_fetch.py
#  Extrae datos macroeconómicos de APIs gratuitas
#  Motor optimizado con Concurrencia y Retries automáticos
# ─────────────────────────────────────────────────────────

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import json
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import os
import concurrent.futures

try:
    from config import FRED_API_KEY, FRED_SERIES, WORLD_BANK_SERIES, WORLD_BANK_COUNTRIES, BANK_META, CATEGORIES
except ImportError:
    raise SystemExit("❌ Falta config.py — asegúrate de estar en la carpeta correcta")

# ─────────────────────────────────────────────────────────
#  Configuración de Sesión Resiliente (Retries & Timeouts)
# ─────────────────────────────────────────────────────────
def get_resilient_session(retries=3, backoff_factor=0.5, status_forcelist=(429, 500, 502, 503, 504)):
    """Crea una sesión HTTP que reintenta automáticamente en caso de fallos temporales."""
    session = requests.Session()
    retry = Retry(
        total=retries,
        read=retries,
        connect=retries,
        backoff_factor=backoff_factor,
        status_forcelist=status_forcelist,
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount('http://', adapter)
    session.mount('https://', adapter)
    return session

http = get_resilient_session()
TIMEOUT = 7  # Segundos máximos de espera por petición

# ─────────────────────────────────────────────────────────
#  Datos fallback (idénticos a los de app.py)
# ─────────────────────────────────────────────────────────
FALLBACK = {
    "FED":  {"currentRate": 3.75, "rates": [{"date":"May-24","r":5.50},{"date":"Jun-24","r":5.50},{"date":"Sep-24","r":5.00},{"date":"Nov-24","r":4.75},{"date":"Dic-24","r":4.50},{"date":"Mar-25","r":4.50},{"date":"Sep-25","r":4.25},{"date":"Oct-25","r":4.00},{"date":"Dic-25","r":3.75},{"date":"Mar-26","r":3.75}], "indicators": [{"cat":"INFLACIÓN","name":"CPI y/y","actual":"3.3%","prev":"2.4%","dev":"+0.9%","dir":1},{"cat":"CRECIMIENTO","name":"Advanced GDP q/q","actual":"0.5%","prev":"1.9%","dev":"-1.4%","dir":-1},{"cat":"EMPLEO","name":"NFP Change","actual":"178K","prev":"-133K","dev":"+311K","dir":1},{"cat":"CONSUMO","name":"Retail Sales m/m","actual":"0.6%","prev":"0.1%","dev":"+0.5%","dir":1}]},
    "BCE":  {"currentRate": 2.15, "rates": [{"date":"Jun-24","r":4.25},{"date":"Sep-24","r":3.65},{"date":"Dic-24","r":3.15},{"date":"Mar-25","r":2.65},{"date":"Jun-25","r":2.15},{"date":"Mar-26","r":2.15}], "indicators": [{"cat":"INFLACIÓN","name":"EZ CPI Flash y/y","actual":"2.6%","prev":"1.9%","dev":"+0.7%","dir":1},{"cat":"CRECIMIENTO","name":"EZ Flash GDP q/q","actual":"0.3%","prev":"0.3%","dev":"0.0%","dir":0},{"cat":"EMPLEO","name":"EZ Unemployment Rate","actual":"6.2%","prev":"6.3%","dev":"-0.1%","dir":1}]},
    "BOE":  {"currentRate": 3.75, "rates": [{"date":"May-24","r":5.25},{"date":"Aug-24","r":5.00},{"date":"Nov-24","r":4.75},{"date":"Feb-25","r":4.50},{"date":"Aug-25","r":4.00},{"date":"Dic-25","r":3.75},{"date":"Mar-26","r":3.75}], "indicators": [{"cat":"INFLACIÓN","name":"CPI y/y","actual":"3.3%","prev":"3.0%","dev":"+0.3%","dir":1},{"cat":"CRECIMIENTO","name":"GDP m/m","actual":"0.2%","prev":"0.3%","dev":"-0.1%","dir":-1},{"cat":"ACTIVIDAD","name":"Flash Manuf. PMI","actual":"51.6","prev":"50.6","dev":"+1.0","dir":1}]},
    "BOC":  {"currentRate": 2.25, "rates": [{"date":"Jun-24","r":4.75},{"date":"Oct-24","r":3.75},{"date":"Dic-24","r":3.25},{"date":"Mar-25","r":2.75},{"date":"Oct-25","r":2.25},{"date":"Mar-26","r":2.25}], "indicators": [{"cat":"INFLACIÓN","name":"CPI m/m","actual":"0.9%","prev":"0.5%","dev":"+0.4%","dir":1},{"cat":"CRECIMIENTO","name":"GDP q/q","actual":"-0.6%","prev":"2.4%","dev":"-3.0%","dir":-1},{"cat":"EMPLEO","name":"Employment Change","actual":"-24.8K","prev":"8.2K","dev":"-33K","dir":-1}]},
    "RBA":  {"currentRate": 4.10, "rates": [{"date":"Jun-24","r":4.35},{"date":"Feb-25","r":4.10},{"date":"May-25","r":3.85},{"date":"Aug-25","r":3.60},{"date":"Feb-26","r":3.85},{"date":"May-26","r":4.10}], "indicators": [{"cat":"INFLACIÓN","name":"CPI y/y","actual":"3.7%","prev":"3.8%","dev":"-0.1%","dir":-1},{"cat":"CRECIMIENTO","name":"GDP q/q","actual":"0.8%","prev":"0.5%","dev":"+0.3%","dir":1}]},
    "RBNZ": {"currentRate": 2.25, "rates": [{"date":"May-24","r":5.50},{"date":"Aug-24","r":5.25},{"date":"Nov-24","r":4.25},{"date":"Feb-25","r":3.75},{"date":"Aug-25","r":3.00},{"date":"Oct-25","r":2.50},{"date":"Apr-26","r":2.25}], "indicators": [{"cat":"INFLACIÓN","name":"CPI q/q","actual":"0.9%","prev":"0.6%","dev":"+0.3%","dir":1},{"cat":"CRECIMIENTO","name":"GDP q/q","actual":"0.2%","prev":"0.9%","dev":"-0.7%","dir":-1}]},
}

# ─────────────────────────────────────────────────────────
#  Helper: leer tasas desde Excel (si existe)
# ─────────────────────────────────────────────────────────
def load_rates_from_excel(excel_path: str, bank_code: str) -> list[dict]:
    if not os.path.exists(excel_path):
        return []
    try:
        wb = load_workbook(excel_path, data_only=True)
        if bank_code not in wb.sheetnames:
            return []
        ws = wb[bank_code]
        rates = []
        for row in range(2, 22):
            fecha = ws.cell(row, 2).value
            tasa  = ws.cell(row, 4).value
            if isinstance(fecha, datetime) and isinstance(tasa, (int, float)):
                rates.append({
                    "date": fecha.strftime("%b-%y"),
                    "r": round(tasa * 100, 2) if tasa < 1 else round(tasa, 2)
                })
        return rates
    except Exception as e:
        print(f"  ⚠ Excel read {bank_code}: {e}")
        return []

def get_current_rate_from_excel(excel_path: str, bank_code: str) -> float | None:
    rates = load_rates_from_excel(excel_path, bank_code)
    if rates:
        return rates[0]["r"]
    return None

# ─────────────────────────────────────────────────────────
#  Funciones de API (optimizadas para hilos)
# ─────────────────────────────────────────────────────────
def _fetch_fred_single(series_id: str, label: str, cat: str) -> dict:
    """Función de trabajo aislada para descargar una serie de la FRED."""
    url = "https://api.stlouisfed.org/fred/series/observations"
    params = {
        "series_id": series_id,
        "api_key": FRED_API_KEY,
        "file_type": "json",
        "sort_order": "desc",
        "limit": 3,
    }
    try:
        r = http.get(url, params=params, timeout=TIMEOUT)
        r.raise_for_status()
        obs = r.json().get("observations", [])
        clean_obs = [{"date": o["date"], "value": o["value"]} for o in obs if o["value"] != "."]
        
        if not clean_obs:
            return {}
            
        latest = float(clean_obs[0]["value"])
        prev = float(clean_obs[1]["value"]) if len(clean_obs) > 1 else latest
        
        if series_id == "FEDFUNDS":
            return {"type": "rate", "value": round(latest, 2)}
            
        dev = round(latest - prev, 4)
        dir_ = 1 if dev > 0 else (-1 if dev < 0 else 0)
        
        if cat in ("INFLACIÓN", "CRECIMIENTO", "CONSUMO"):
            fmt_v, fmt_p, fmt_d = f"{latest:.1%}", f"{prev:.1%}", f"{'+' if dev > 0 else ''}{dev:.1%}"
        else:
            fmt_v, fmt_p, fmt_d = f"{latest:,.1f}", f"{prev:,.1f}", f"{'+' if dev > 0 else ''}{dev:,.1f}"
            
        return {
            "type": "indicator",
            "name": label,
            "data": {
                "cat": cat, "name": label, "actual": fmt_v, "prev": fmt_p,
                "dev": fmt_d, "dir": dir_, "raw_actual": latest, "raw_prev": prev,
                "date": clean_obs[0]["date"]
            }
        }
    except Exception as e:
        return {}

def build_fed_indicators() -> tuple[dict, float]:
    indicators = {}
    rate = None
    
    # Ejecución paralela de hasta 5 peticiones simultáneas
    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        futures = [executor.submit(_fetch_fred_single, s_id, label, cat) 
                   for s_id, (label, cat) in FRED_SERIES.items()]
        
        for future in concurrent.futures.as_completed(futures):
            res = future.result()
            if res.get("type") == "rate":
                rate = res["value"]
            elif res.get("type") == "indicator":
                indicators[res["name"]] = res["data"]

    return indicators, rate

def fetch_ecb_rate() -> float | None:
    url = "https://data-api.ecb.europa.eu/service/data/ICP/M.U2.EUR.4F.DP00?format=jsondata&lastNObservations=1"
    try:
        r = http.get(url, timeout=TIMEOUT)
        r.raise_for_status()
        data = r.json()
        obs = data.get('dataSets', [{}])[0].get('series', {}).get('0:0:0:0:0', {}).get('observations', {})
        if not obs:
            obs = data.get('observations', {})
        if obs:
            last_key = max(obs.keys(), key=int)
            return round(float(obs[last_key][0]), 4)
    except Exception:
        pass
    return None

def _fetch_wb_single(country: str, series_id: str, label: str, cat: str) -> dict:
    """Función de trabajo aislada para el Banco Mundial."""
    url = f"https://api.worldbank.org/v2/country/{country}/indicator/{series_id}"
    params = {"format": "json", "per_page": 5, "mrv": 5}
    try:
        r = http.get(url, params=params, timeout=TIMEOUT)
        r.raise_for_status()
        data = r.json()
        if len(data) < 2 or not data[1]:
            return {}
            
        clean_obs = [{"date": d["date"], "value": d["value"]} for d in data[1] if d["value"] is not None]
        if not clean_obs:
            return {}
            
        latest = clean_obs[0]["value"]
        prev = clean_obs[1]["value"] if len(clean_obs) > 1 else latest
        dev = round(latest - prev, 4)
        dir_ = 1 if dev > 0 else (-1 if dev < 0 else 0)
        
        return {
            "name": label,
            "data": {
                "cat": cat, "name": label, "actual": f"{latest:.2f}%", "prev": f"{prev:.2f}%",
                "dev": f"{'+' if dev > 0 else ''}{dev:.2f}%", "dir": dir_,
                "raw_actual": latest, "raw_prev": prev, "date": clean_obs[0]["date"]
            }
        }
    except Exception:
        return {}

def build_wb_indicators(bank_code: str) -> dict:
    country = WORLD_BANK_COUNTRIES.get(bank_code)
    if not country:
        return {}
    
    indicators = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        futures = [executor.submit(_fetch_wb_single, country, s_id, label, cat) 
                   for s_id, (label, cat) in WORLD_BANK_SERIES.items()]
        
        for future in concurrent.futures.as_completed(futures):
            res = future.result()
            if res:
                indicators[res["name"]] = res["data"]
                
    return indicators

# ─────────────────────────────────────────────────────────
#  Sentimiento (reglas simples)
# ─────────────────────────────────────────────────────────
CAT_BULLISH_WHEN_UP = {"CRECIMIENTO", "EMPLEO", "CONSUMO", "ACTIVIDAD"}
CAT_BULLISH_WHEN_DOWN = {"INFLACIÓN"}

def calc_sentiment(indicators: dict, category: str) -> str:
    relevant = [v for v in indicators.values() if v.get("cat") == category]
    if not relevant:
        return "NEUTRO"
    score = 0
    for ind in relevant:
        d = ind.get("dir", 0)
        if category in CAT_BULLISH_WHEN_UP:
            score += d
        elif category in CAT_BULLISH_WHEN_DOWN:
            score -= d
    if score > 0:
        return "BULLISH"
    elif score < 0:
        return "BEARISH"
    return "NEUTRO"

# ─────────────────────────────────────────────────────────
#  Fetch principal
# ─────────────────────────────────────────────────────────
def fetch_all(excel_path: str = "Datos_Macro1.xlsm") -> dict:
    print(f"\n{'='*55}")
    print(f"  MacroVision · Sincronización Automática {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print(f"{'='*55}")

    result = {}
    for bank_code, meta in BANK_META.items():
        print(f"  [{bank_code}] Obteniendo datos (API/Excel)...")
        indicators = {}
        rate = None

        if bank_code == "FED" and FRED_API_KEY and FRED_API_KEY != "5b53bdf5fe99ed21dc2af47440603b37":
            indicators, rate = build_fed_indicators()
        elif bank_code == "BCE":
            rate = fetch_ecb_rate()
            indicators = build_wb_indicators(bank_code)
        else:
            indicators = build_wb_indicators(bank_code)

        if rate is None:
            rate = get_current_rate_from_excel(excel_path, bank_code)
        if rate is None:
            rate = FALLBACK.get(bank_code, {}).get("currentRate", 0.0)

        if not indicators:
            fallback_indicators = FALLBACK.get(bank_code, {}).get("indicators", [])
            for ind in fallback_indicators:
                indicators[ind["name"]] = ind

        sentiment = {cat: calc_sentiment(indicators, cat) for cat in CATEGORIES}

        rates_hist = load_rates_from_excel(excel_path, bank_code)
        if not rates_hist:
            rates_hist = FALLBACK.get(bank_code, {}).get("rates", [])

        if rate and (not rates_hist or rates_hist[0].get("r") != rate):
            today_label = datetime.today().strftime("%b-%y")
            rates_hist.insert(0, {"date": today_label, "r": rate})

        result[bank_code] = {
            "name": bank_code,
            "flag": meta["flag"],
            "currency": meta["currency"],
            "fullName": meta["name"],
            "currentRate": rate,
            "lastMeeting": datetime.today().strftime("%d %b %Y"),
            "rates": rates_hist,
            "sentiment": sentiment,
            "indicators": list(indicators.values()),
            "updated": datetime.now().isoformat(),
        }
        print(f"    ↳ Tasa: {rate}% | Indicadores actualizados: {len(indicators)}")

    return result

def save_json(data: dict, path: str = "macro_data.json"):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"\n  💾 Datos guardados exitosamente → {os.path.abspath(path)}")

if __name__ == "__main__":
    data = fetch_all()
    save_json(data)
    print("\n  ✅ Proceso finalizado. El dashboard de Streamlit se actualizará automáticamente.\n")